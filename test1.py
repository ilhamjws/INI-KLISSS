import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Pengeluaran Uang Operasional")
        self.root.geometry("400x300")

        # Inisialisasi daftar pengeluaran
        self.expenses = []

        # Nama file Excel
        self.filename = "pengeluaran_operasional.xlsx"
        
        # Buat file Excel jika belum ada
        self.create_excel_file()

        # Label dan Entry untuk deskripsi pengeluaran
        tk.Label(root, text="Deskripsi Pengeluaran").grid(row=0, column=0, padx=10, pady=10)
        self.desc_entry = tk.Entry(root, width=30)
        self.desc_entry.grid(row=0, column=1, padx=10, pady=10)

        # Label dan Entry untuk jumlah pengeluaran
        tk.Label(root, text="Jumlah (Rp)").grid(row=1, column=0, padx=10, pady=10)
        self.amount_entry = tk.Entry(root, width=30)
        self.amount_entry.grid(row=1, column=1, padx=10, pady=10)

        # Tombol untuk menambah pengeluaran
        tk.Button(root, text="Tambah Pengeluaran", command=self.add_expense).grid(row=2, column=0, columnspan=2, pady=10)

        # Tampilkan total pengeluaran
        self.total_label = tk.Label(root, text="Total Pengeluaran: Rp 0")
        self.total_label.grid(row=3, column=0, columnspan=2, pady=10)

        # Tombol untuk reset pengeluaran
        tk.Button(root, text="Reset Pengeluaran", command=self.reset_expenses).grid(row=4, column=0, columnspan=2, pady=10)

    def create_excel_file(self):
        # Jika file belum ada, buat file baru dengan header
        if not os.path.exists(self.filename):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Pengeluaran"
            sheet.append(["Deskripsi", "Jumlah (Rp)"])
            workbook.save(self.filename)

    def add_expense(self):
        try:
            # Ambil data dari Entry
            description = self.desc_entry.get()
            amount = float(self.amount_entry.get())
            
            # Tambahkan pengeluaran ke dalam daftar
            self.expenses.append((description, amount))

            # Bersihkan Entry
            self.desc_entry.delete(0, tk.END)
            self.amount_entry.delete(0, tk.END)

            # Perbarui total pengeluaran
            self.update_total()

            # Simpan ke file Excel
            self.save_to_excel(description, amount)

            messagebox.showinfo("Info", "Pengeluaran berhasil ditambahkan dan disimpan ke Excel!")
        except ValueError:
            messagebox.showerror("Error", "Masukkan jumlah yang valid!")

    def update_total(self):
        total = sum(amount for _, amount in self.expenses)
        self.total_label.config(text=f"Total Pengeluaran: Rp {total:,.2f}")

    def save_to_excel(self, description, amount):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook["Pengeluaran"]
        sheet.append([description, amount])
        workbook.save(self.filename)

    def reset_expenses(self):
        self.expenses.clear()
        self.update_total()
        # Kosongkan file Excel, simpan hanya header
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Pengeluaran"
        sheet.append(["Deskripsi", "Jumlah (Rp)"])
        workbook.save(self.filename)
        messagebox.showinfo("Info", "Semua pengeluaran telah direset dan data Excel dikosongkan!")

# Main program
if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
