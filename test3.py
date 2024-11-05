import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
import os
import matplotlib.pyplot as plt

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Pengeluaran Uang Operasional")
        self.root.geometry("500x700")
        self.root.resizable(False, False)

        # Menggunakan tema ttk untuk tampilan lebih baik
        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", background="#007ACC")
        style.configure("TLabel", font=("Arial", 10))
        
        # Menambahkan logo ke GUI
        self.add_logo(root)

        # Judul utama
        title_label = ttk.Label(root, text="Aplikasi Pengeluaran Operasional", font=("Arial", 16, "bold"), foreground="#007ACC")
        title_label.pack(pady=10)

        # Frame untuk formulir input pengeluaran
        input_frame = ttk.Frame(root, padding="10")
        input_frame.pack(pady=10, fill="x", padx=20)

        # Label dan combobox untuk memilih bulan
        ttk.Label(input_frame, text="Pilih Bulan", foreground="#007ACC").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.month_var = tk.StringVar()
        self.month_combobox = ttk.Combobox(input_frame, textvariable=self.month_var, values=[
            "Januari", "Februari", "Maret", "April", "Mei", "Juni", 
            "Juli", "Agustus", "September", "Oktober", "November", "Desember"
        ])
        self.month_combobox.grid(row=0, column=1, sticky="ew", padx=5, pady=5)

        # Label dan entry untuk deskripsi pengeluaran
        ttk.Label(input_frame, text="Deskripsi Pengeluaran", foreground="#007ACC").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.desc_entry = ttk.Entry(input_frame, width=30)
        self.desc_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        # Label dan entry untuk harga satuan
        ttk.Label(input_frame, text="Harga Satuan (Rp)", foreground="#007ACC").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.unit_price_entry = ttk.Entry(input_frame, width=30)
        self.unit_price_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

        # Label dan entry untuk kuantitas
        ttk.Label(input_frame, text="Kuantitas", foreground="#007ACC").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.quantity_entry = ttk.Entry(input_frame, width=30)
        self.quantity_entry.grid(row=3, column=1, sticky="ew", padx=5, pady=5)

        # Garis pemisah
        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=20, pady=10)

        # Frame untuk tombol aksi
        button_frame = ttk.Frame(root, padding="10")
        button_frame.pack(pady=10, fill="x", padx=20)

        # Tombol untuk menambah pengeluaran
        add_button = ttk.Button(button_frame, text="Tambah Pengeluaran", command=self.add_expense)
        add_button.pack(pady=5, fill="x")

        # Tombol untuk reset pengeluaran bulan
        reset_button = ttk.Button(button_frame, text="Reset Pengeluaran Bulan", command=self.reset_expenses)
        reset_button.pack(pady=5, fill="x")

        # Tombol untuk menampilkan grafik pie chart
        pie_chart_button = ttk.Button(button_frame, text="Tampilkan Pie Chart", command=self.show_pie_chart)
        pie_chart_button.pack(pady=5, fill="x")

        # Tombol untuk menghapus pengeluaran yang dipilih
        delete_button = ttk.Button(button_frame, text="Hapus Pengeluaran Terpilih", command=self.delete_expense)
        delete_button.pack(pady=5, fill="x")

        # Garis pemisah
        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=20, pady=10)

        # Tampilkan total pengeluaran
        self.total_label = ttk.Label(root, text="Total Pengeluaran: Rp 0", font=("Arial", 12, "bold"), foreground="#007ACC")
        self.total_label.pack(pady=10)

        # Frame untuk daftar pengeluaran
        expense_list_frame = ttk.Frame(root, padding="10")
        expense_list_frame.pack(pady=10, fill="both", expand=True, padx=20)

        # Tabel daftar pengeluaran
        columns = ("Deskripsi", "Harga Satuan (Rp)", "Kuantitas", "Jumlah Total (Rp)")
        self.expense_tree = ttk.Treeview(expense_list_frame, columns=columns, show="headings")
        self.expense_tree.heading("Deskripsi", text="Deskripsi")
        self.expense_tree.heading("Harga Satuan (Rp)", text="Harga Satuan (Rp)")
        self.expense_tree.heading("Kuantitas", text="Kuantitas")
        self.expense_tree.heading("Jumlah Total (Rp)", text="Jumlah Total (Rp)")
        self.expense_tree.pack(fill="both", expand=True)

        # Inisialisasi daftar pengeluaran
        self.expenses = []

        # Nama file Excel
        self.filename = "pengeluaran_operasional_bulanan.xlsx"
        
        # Buat file Excel jika belum ada
        self.create_excel_file()

    def add_logo(self, root):
        image = Image.open("WhatsApp Image 2024-10-01 at 09.41.57_a1a67c14.jpg")
        image = image.resize((100, 100), Image.LANCZOS)
        self.logo_image = ImageTk.PhotoImage(image)
        logo_label = tk.Label(root, image=self.logo_image)
        logo_label.pack(pady=5)

    def create_excel_file(self):
        if not os.path.exists(self.filename):
            workbook = Workbook()
            workbook.save(self.filename)

    def add_expense(self):
        try:
            month = self.month_var.get()
            description = self.desc_entry.get()
            unit_price = float(self.unit_price_entry.get())
            quantity = int(self.quantity_entry.get())
            total_amount = unit_price * quantity
            
            if not month:
                messagebox.showerror("Error", "Pilih bulan terlebih dahulu!")
                return

            self.desc_entry.delete(0, tk.END)
            self.unit_price_entry.delete(0, tk.END)
            self.quantity_entry.delete(0, tk.END)

            # Tambahkan pengeluaran ke daftar
            self.expenses.append((description, unit_price, quantity, total_amount))

            # Tambahkan ke Treeview daftar pengeluaran
            self.expense_tree.insert("", "end", values=(description, unit_price, quantity, total_amount))

            # Perbarui total pengeluaran
            self.update_total()

            # Simpan ke file Excel
            self.save_to_excel(month, description, unit_price, quantity, total_amount)

            messagebox.showinfo("Info", "Pengeluaran berhasil ditambahkan dan disimpan ke Excel!")
        except ValueError:
            messagebox.showerror("Error", "Masukkan jumlah dan harga satuan yang valid!")

    def update_total(self):
        total = sum(expense[3] for expense in self.expenses)  # jumlah total dari kolom "Jumlah Total"
        self.total_label.config(text=f"Total Pengeluaran: Rp {total:,.2f}")

    def save_to_excel(self, month, description, unit_price, quantity, total_amount):
        # Memuat workbook dari file Excel
        workbook = openpyxl.load_workbook(self.filename)

        # Periksa apakah lembar untuk bulan tersebut ada; jika tidak, buat lembar baru
        if month not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=month)
            # Tambahkan header kolom ke lembar baru
            sheet.append(["Deskripsi", "Harga Satuan (Rp)", "Kuantitas", "Jumlah Total (Rp)"])
        else:
            sheet = workbook[month]
        
        # Tambahkan data ke lembar kerja
        sheet.append([description, unit_price, quantity, total_amount])

        # Simpan workbook ke file Excel
        workbook.save(self.filename)
        print(f"Data berhasil disimpan di lembar '{month}'")

    def delete_expense(self):
        selected_item = self.expense_tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Pilih pengeluaran yang ingin dihapus!")
            return

        for item in selected_item:
            values = self.expense_tree.item(item, "values")
            description = values[0]

            # Hapus dari daftar pengeluaran
            self.expenses = [expense for expense in self.expenses if expense[0] != description]

            # Hapus dari Treeview
            self.expense_tree.delete(item)

        # Perbarui total pengeluaran
        self.update_total()
        messagebox.showinfo("Info", "Pengeluaran terpilih berhasil dihapus!")

    def reset_expenses(self):
        self.expenses.clear()
        for item in self.expense_tree.get_children():
            self.expense_tree.delete(item)
        self.update_total()
        messagebox.showinfo("Info", "Semua pengeluaran bulan ini berhasil direset!")

    def show_pie_chart(self):
        if not self.expenses:
            messagebox.showerror("Error", "Tidak ada pengeluaran untuk ditampilkan dalam pie chart!")
            return

        labels = [expense[0] for expense in self.expenses]
        amounts = [expense[3] for expense in self.expenses]
        plt.figure(figsize=(6, 6))
        plt.pie(amounts, labels=labels, autopct="%1.1f%%", startangle=90)
        plt.title("Distribusi Pengeluaran")
        plt.show()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
